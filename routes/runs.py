import json
from datetime import datetime

from flask import flash, jsonify, redirect, render_template, request, send_file, url_for

from config import (
    add_default_tag,
    db_log,
    get_backup_locations,
    get_current_user,
    get_default_tags,
    get_slack_enabled,
    load_crispr_libraries,
    load_run_templates,
    load_workflows,
)
from markupsafe import Markup

from helpers import (
    _compare_configs,
    _parse_datetime,
    extract_samples_from_config,
    extract_samples_from_sample_info,
    find_all_duplicate_run_ids,
    find_duplicate_runs,
    get_journal_name,
)
from models import (
    FILE_TYPES,
    RUN_STATUSES,
    AttachedFile,
    Project,
    ProjectScript,
    ResearchGroup,
    Researcher,
    WorkflowRun,
    db,
)


def register(app):
    PER_PAGE = 50

    @app.route("/runs/parse-xlsx", methods=["POST"])
    def parse_run_xlsx():
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file uploaded"}), 400
        try:
            import openpyxl
            from datetime import date, datetime as dt
        except ImportError:
            return jsonify({"error": "openpyxl not installed"}), 500
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            return jsonify({"error": f"Cannot open file: {e}"}), 400
        if "Run Template" not in wb.sheetnames:
            return jsonify({"error": "File has no 'Run Template' sheet — is this a valid NGS Tracker template?"}), 400
        ws = wb["Run Template"]

        # Scan column A for field labels so the parser works regardless of
        # which template version was used (row numbers may differ).
        fields: dict[str, object] = {}
        tags: dict[str, bool] = {}
        in_tags = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            label_cell = row[0]
            value_cell = row[1] if len(row) > 1 else None
            label = str(label_cell.value or "").strip()
            raw = value_cell.value if value_cell else None

            def _val():
                if raw is None:
                    return ""
                if isinstance(raw, (date, dt)):
                    return raw.strftime("%Y-%m-%d")
                return str(raw).strip()

            if in_tags:
                if label and not label.lower().startswith("tags"):
                    tags[label] = str(raw or "").strip().lower() == "yes"
                continue

            low = label.lower()
            if low.startswith("workflow") and "tag" not in low and "version" not in low:
                fields["workflow"] = _val()
            elif low.startswith("version") or low.startswith("release"):
                fields["version"] = _val()
            elif low.startswith("run date") or low.startswith("date"):
                fields["run_date"] = _val()
            elif low == "status":
                fields["status"] = _val()
            elif low.startswith("short description") or low.startswith("description"):
                fields["description"] = _val()
            elif low == "notes":
                fields["notes"] = _val()
            elif low.startswith("tags"):
                in_tags = True

        valid_statuses = {"completed", "running", "failed", "pending"}
        status = fields.get("status", "").lower()
        if status not in valid_statuses:
            status = "completed"

        return jsonify({
            "workflow":    fields.get("workflow", ""),
            "version":     fields.get("version", ""),
            "run_date":    fields.get("run_date", ""),
            "status":      status,
            "description": fields.get("description", ""),
            "notes":       fields.get("notes", ""),
            "tags":        tags,
        })

    @app.route("/runs")
    def runs_list():
        sort = request.args.get("sort", "date")
        direction = request.args.get("dir", "desc")
        tag_filter = request.args.get("tag", "").strip()
        tag_filters = [t.strip() for t in tag_filter.split(",") if t.strip()]
        tag_mode = request.args.get("tag_mode", "or")
        status_filter = request.args.get("status", "").strip().lower()
        workflow_filter = request.args.get("workflow", "").strip()
        library_filter = request.args.get("library", "").strip()
        low_mapping_filter = bool(request.args.get("low_mapping", ""))
        duplicates_filter = bool(request.args.get("duplicates", ""))
        date_from_str = request.args.get("date_from", "").strip()
        date_to_str = request.args.get("date_to", "").strip()
        journal_filter = request.args.get("journal", "").strip()
        page = max(1, request.args.get("page", 1, type=int))
        reverse = direction == "desc"

        # Parse and validate date range
        date_from = None
        date_to = None
        try:
            if date_from_str:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            date_from_str = ""
        try:
            if date_to_str:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            date_to_str = ""

        all_runs = WorkflowRun.query.filter_by(trashed=False).all()
        all_tags = sorted({tag for r in all_runs for tag in r.tag_list})
        all_workflows = sorted({r.workflow_name for r in all_runs})

        # Runs with at least one sample below the workflow's mapping rate cutoff
        _wf_cutoffs = {
            w["name"]: float(w.get("mapping_rate_cutoff", 60.0))
            for w in load_workflows()
        }

        def _has_low_mapping(run):
            cutoff = _wf_cutoffs.get(run.workflow_name, 60.0)
            for f in run.attached_files:
                if f.file_type == "mapping_rates" and f.config_dict:
                    if any(r < cutoff for r in f.config_dict.get("rates", [])):
                        return True
            return False

        low_mapping_run_ids = {r.id for r in all_runs if _has_low_mapping(r)}
        duplicate_run_ids = find_all_duplicate_run_ids(all_runs)

        runs = all_runs
        if workflow_filter:
            runs = [r for r in runs if r.workflow_name == workflow_filter]
        crispr_libraries_in_runs = sorted({
            r.crispr_library for r in runs
            if r.workflow_name == "crispr-screens" and r.crispr_library
        })
        if library_filter and workflow_filter == "crispr-screens":
            runs = [r for r in runs if r.crispr_library == library_filter]
        if low_mapping_filter:
            runs = [r for r in runs if r.id in low_mapping_run_ids]
        if duplicates_filter:
            runs = [r for r in runs if r.id in duplicate_run_ids]
        if journal_filter:
            pub_project_ids = {
                p.id
                for p in Project.query.filter(
                    Project.trashed == False,
                    Project.published == True,
                    Project.publication_url.isnot(None),
                    Project.publication_url != "",
                ).all()
                if get_journal_name(p.publication_url) == journal_filter
            }
            runs = [r for r in runs if r.project_id in pub_project_ids]
        if tag_filters:
            if tag_mode == "and":
                runs = [r for r in runs if all(t in r.tag_list for t in tag_filters)]
            else:
                runs = [r for r in runs if any(t in r.tag_list for t in tag_filters)]
        if status_filter and status_filter in RUN_STATUSES:
            runs = [r for r in runs if r.status == status_filter]
        if date_from:
            runs = [r for r in runs if r.run_date.date() >= date_from]
        if date_to:
            runs = [r for r in runs if r.run_date.date() <= date_to]

        key_map = {
            "workflow": lambda r: r.workflow_name.lower(),
            "project": lambda r: r.project.name.lower(),
            "researcher": lambda r: r.project.researcher.name.lower(),
            "date": lambda r: r.run_date,
            "status": lambda r: r.status,
            "backup": lambda r: len(r.backup_labels),
            "files": lambda r: len(r.attached_files),
        }
        runs.sort(key=key_map.get(sort, key_map["date"]), reverse=reverse)

        total = len(runs)
        total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
        page = min(page, total_pages)
        runs_page = runs[(page - 1) * PER_PAGE : page * PER_PAGE]

        scripts = (
            ProjectScript.query.filter_by(trashed=False)
            .order_by(ProjectScript.uploaded_at.desc())
            .all()
        )

        return render_template(
            "runs/list.html",
            runs=runs_page,
            sort=sort,
            dir=direction,
            tag_filter=tag_filter,
            tag_filters=tag_filters,
            tag_mode=tag_mode,
            status_filter=status_filter,
            workflow_filter=workflow_filter,
            library_filter=library_filter,
            crispr_libraries_in_runs=crispr_libraries_in_runs,
            low_mapping_filter=low_mapping_filter,
            low_mapping_run_ids=low_mapping_run_ids,
            duplicates_filter=duplicates_filter,
            duplicate_run_ids=duplicate_run_ids,
            date_from=date_from_str,
            date_to=date_to_str,
            journal_filter=journal_filter,
            all_tags=all_tags,
            all_workflows=all_workflows,
            run_statuses=RUN_STATUSES,
            page=page,
            total_pages=total_pages,
            total=total,
            per_page=PER_PAGE,
            backup_locations=get_backup_locations(),
            scripts=scripts,
        )

    @app.route("/runs/<int:id>")
    def run_detail(id):
        run = db.get_or_404(WorkflowRun, id)
        workflows = load_workflows()
        wf_urls = {w["name"]: w.get("url") or None for w in workflows}
        wf_entry = next((w for w in workflows if w["name"] == run.workflow_name), {})
        mapping_rate_cutoff = float(wf_entry.get("mapping_rate_cutoff", 60.0))

        # Runs with at least one parsed config, excluding this run, for the compare picker
        config_run_ids = (
            db.session.query(AttachedFile.workflow_run_id)
            .filter(
                AttachedFile.file_type == "config",
                AttachedFile.parsed_config.isnot(None),
            )
            .distinct()
            .subquery()
        )
        compare_runs = (
            WorkflowRun.query.filter(
                WorkflowRun.id != id,
                WorkflowRun.trashed == False,
                WorkflowRun.id.in_(config_run_ids),
            )
            .order_by(WorkflowRun.workflow_name, WorkflowRun.run_date.desc())
            .all()
        )

        backup_locations = get_backup_locations()
        backed_up = {b["location"]: b["path"] for b in run.backups_list}
        loc_names = [l["name"] for l in backup_locations]
        duplicate_runs = find_duplicate_runs(
            run.project_id,
            run.workflow_name,
            run.run_date,
            exclude_id=run.id,
            description=run.description or "",
            sample_names={rs.sample.name for rs in run.run_samples},
        )

        config_samples = None
        for f in run.attached_files:
            if f.file_type == "config" and f.config_dict:
                extracted = extract_samples_from_config(
                    run.workflow_name, f.config_dict
                )
                if extracted:
                    config_samples = extracted
                    break

        has_sample_info = False
        for f in run.attached_files:
            if f.file_type == "sample_info":
                from config import resolve_stored_path
                stored = resolve_stored_path(f.stored_path)
                if stored.exists():
                    result = extract_samples_from_sample_info(run.workflow_name, stored)
                    if result:
                        has_sample_info = True
                        break

        return render_template(
            "runs/detail.html",
            run=run,
            file_types=FILE_TYPES,
            wf_urls=wf_urls,
            compare_runs=compare_runs,
            backup_locations=backup_locations,
            backed_up=backed_up,
            loc_names=loc_names,
            mapping_rate_cutoff=mapping_rate_cutoff,
            duplicate_runs=duplicate_runs,
            config_samples=config_samples,
            has_sample_info=has_sample_info,
            slack_enabled=get_slack_enabled(),
            crispr_libraries=load_crispr_libraries(),
        )

    @app.route("/runs/new", methods=["GET", "POST"])
    def run_new():
        projects = (
            Project.query.filter_by(trashed=False)
            .join(Researcher)
            .filter(Researcher.trashed == False)
            .join(ResearchGroup)
            .filter(ResearchGroup.trashed == False)
            .order_by(ResearchGroup.name, Researcher.name, Project.name)
            .all()
        )
        if not projects:
            flash("Create a project first.", "warning")
            return redirect(url_for("project_new"))

        preselected = request.args.get("project_id", type=int)
        template_id = request.args.get("template_id", "")
        prefill = next(
            (t for t in load_run_templates() if t.get("id") == template_id),
            None,
        )

        if request.method == "POST":
            project_id = request.form.get("project_id", type=int)
            workflow_name = request.form.get("workflow_name", "").strip()
            workflow_tag = request.form.get("workflow_tag", "").strip()
            description = request.form.get("description", "").strip()
            notes = request.form.get("notes", "").strip()
            shared_storage_path = request.form.get("shared_storage_path", "").strip()
            status = request.form.get("status", "completed")
            run_date = _parse_datetime(request.form.get("run_date", ""))
            selected_locs = set(request.form.getlist("backup_loc"))
            backups = []
            for loc in get_backup_locations():
                if loc["name"] not in selected_locs:
                    continue
                subpath = request.form.get(f"backup_path_{loc['name']}", "").strip()
                base = loc.get("base_path", "").strip()
                if base and subpath:
                    full = base.rstrip("/") + "/" + subpath.lstrip("/")
                elif base:
                    full = base
                else:
                    full = subpath
                backups.append({"location": loc["name"], "path": full})
            wf_list = load_workflows()
            wf_systems = {w["name"]: w.get("system", "snakemake") for w in wf_list}
            workflow_system = wf_systems.get(workflow_name, "other")

            # Persist any newly created tags to the defaults list
            for t in request.form.get("new_tags", "").split(","):
                add_default_tag(t.strip())
            selected_tags = request.form.getlist("tags")
            tags = ", ".join(t for t in selected_tags if t.strip())

            if not workflow_name or not project_id:
                flash("Workflow name and project are required.", "danger")
                default_tags = get_default_tags()
                return render_template(
                    "runs/form.html",
                    run=None,
                    projects=projects,
                    preselected=preselected,
                    file_types=FILE_TYPES,
                    workflows=wf_list,
                    run_statuses=RUN_STATUSES,
                    default_tags=default_tags,
                    extra_tags=[],
                    backup_locations=get_backup_locations(),
                    prefill=prefill,
                    templates=load_run_templates(),
                    wf_systems_json=json.dumps(wf_systems),
                )

            crispr_library = request.form.get("crispr_library", "").strip()

            run = WorkflowRun(
                project_id=project_id,
                workflow_name=workflow_name,
                workflow_tag=workflow_tag,
                description=description,
                tags=tags,
                run_date=run_date,
                notes=notes,
                status=status,
                backups=json.dumps(backups),
                shared_storage_path=shared_storage_path,
                workflow_system=workflow_system,
                created_by=get_current_user(),
                crispr_library=crispr_library,
            )
            db.session.add(run)
            db.session.commit()
            db_log(
                "CREATE",
                "WorkflowRun",
                run.id,
                f"{run.workflow_name} (project: {run.project.name})",
            )
            flash(f'Workflow run "{workflow_name}" created.', "success")
            duplicates = find_duplicate_runs(
                project_id,
                workflow_name,
                run_date,
                exclude_id=run.id,
                description=run.description or "",
                sample_names={rs.sample.name for rs in run.run_samples},
            )
            if duplicates:
                links = ", ".join(
                    f'<a href="{url_for("run_detail", id=r.id)}" class="alert-link">'
                    f"{r.workflow_name} #{r.id} ({r.run_date.strftime('%Y-%m-%d')})</a>"
                    for r in duplicates
                )
                flash(
                    Markup(
                        f"Possible duplicate: another <strong>{workflow_name}</strong> run "
                        f"exists for this project on the same date — {links}. "
                        f"Delete one if this was accidental."
                    ),
                    "warning",
                )
            return redirect(url_for("run_detail", id=run.id))

        default_tags = get_default_tags()
        wf_list = load_workflows()
        return render_template(
            "runs/form.html",
            run=None,
            projects=projects,
            preselected=preselected,
            file_types=FILE_TYPES,
            workflows=wf_list,
            run_statuses=RUN_STATUSES,
            default_tags=default_tags,
            extra_tags=[],
            backup_locations=get_backup_locations(),
            prefill=prefill,
            templates=load_run_templates(),
            wf_systems_json=json.dumps(
                {w["name"]: w.get("system", "snakemake") for w in wf_list}
            ),
            crispr_libraries=load_crispr_libraries(),
        )

    @app.route("/runs/<int:id>/edit", methods=["GET", "POST"])
    def run_edit(id):
        run = db.get_or_404(WorkflowRun, id)
        projects = (
            Project.query.filter_by(trashed=False)
            .join(Researcher)
            .filter(Researcher.trashed == False)
            .join(ResearchGroup)
            .filter(ResearchGroup.trashed == False)
            .order_by(ResearchGroup.name, Researcher.name, Project.name)
            .all()
        )
        if request.method == "POST":
            run.workflow_name = request.form.get("workflow_name", "").strip()
            run.workflow_tag = request.form.get("workflow_tag", "").strip()
            run.description = request.form.get("description", "").strip()
            run.notes = request.form.get("notes", "").strip()
            run.shared_storage_path = request.form.get(
                "shared_storage_path", ""
            ).strip()
            run.status = request.form.get("status", "completed")
            run.run_date = _parse_datetime(request.form.get("run_date", ""))
            selected_locs = set(request.form.getlist("backup_loc"))
            edit_backups = []
            for loc in get_backup_locations():
                if loc["name"] not in selected_locs:
                    continue
                subpath = request.form.get(f"backup_path_{loc['name']}", "").strip()
                base = loc.get("base_path", "").strip()
                if base and subpath:
                    full = base.rstrip("/") + "/" + subpath.lstrip("/")
                elif base:
                    full = base
                else:
                    full = subpath
                edit_backups.append({"location": loc["name"], "path": full})
            run.backups = json.dumps(edit_backups)
            wf_systems = {
                w["name"]: w.get("system", "snakemake") for w in load_workflows()
            }
            run.workflow_system = wf_systems.get(run.workflow_name, "other")
            run.crispr_library = request.form.get("crispr_library", "").strip()
            for t in request.form.get("new_tags", "").split(","):
                add_default_tag(t.strip())
            selected_tags = request.form.getlist("tags")
            run.tags = ", ".join(t for t in selected_tags if t.strip())
            db.session.commit()
            db_log(
                "UPDATE", "WorkflowRun", id, f"{run.workflow_name} status={run.status}"
            )
            flash("Workflow run updated.", "success")
            return redirect(url_for("run_detail", id=id))
        default_tags = get_default_tags()
        default_tag_names = {t["name"] for t in default_tags}
        extra_tags = [t for t in run.tag_list if t not in default_tag_names]
        wf_list = load_workflows()
        return render_template(
            "runs/form.html",
            run=run,
            projects=projects,
            preselected=run.project_id,
            file_types=FILE_TYPES,
            workflows=wf_list,
            run_statuses=RUN_STATUSES,
            default_tags=default_tags,
            extra_tags=extra_tags,
            backup_locations=get_backup_locations(),
            prefill=None,
            templates=[],
            wf_systems_json=json.dumps(
                {w["name"]: w.get("system", "snakemake") for w in wf_list}
            ),
            crispr_libraries=load_crispr_libraries(),
        )

    @app.route("/runs/<int:id>/clone", methods=["POST"])
    def run_clone(id):
        src = db.get_or_404(WorkflowRun, id)
        clone = WorkflowRun(
            project_id=src.project_id,
            workflow_name=src.workflow_name,
            workflow_tag=src.workflow_tag,
            description=src.description,
            tags=src.tags,
            notes=src.notes,
            status="pending",
            backups=json.dumps(src.backups_list),
            workflow_system=src.workflow_system or "snakemake",
        )
        db.session.add(clone)
        db.session.commit()
        db_log(
            "CREATE",
            "WorkflowRun",
            clone.id,
            f"{clone.workflow_name} (cloned from id={src.id})",
        )
        flash(f'Run cloned from "{src.workflow_name}" — review and save.', "success")
        return redirect(url_for("run_edit", id=clone.id))

    @app.route("/runs/compare")
    def run_compare():
        id_a = request.args.get("a", type=int)
        id_b = request.args.get("b", type=int)
        if not id_a or not id_b:
            flash("Two run IDs are required for comparison.", "danger")
            return redirect(url_for("runs_list"))

        run_a = db.get_or_404(WorkflowRun, id_a)
        run_b = db.get_or_404(WorkflowRun, id_b)

        config_a = next(
            (
                f.config_dict
                for f in run_a.attached_files
                if f.file_type == "config" and f.config_dict
            ),
            None,
        )
        config_b = next(
            (
                f.config_dict
                for f in run_b.attached_files
                if f.file_type == "config" and f.config_dict
            ),
            None,
        )

        if not config_a and not config_b:
            flash("Neither run has a parsed Snakemake config.", "warning")
            return redirect(url_for("run_detail", id=id_a))

        diff_rows = _compare_configs(config_a or {}, config_b or {})
        counts = {
            s: sum(1 for r in diff_rows if r["status"] == s)
            for s in ("same", "changed", "added", "removed")
        }

        return render_template(
            "runs/compare.html",
            run_a=run_a,
            run_b=run_b,
            diff_rows=diff_rows,
            counts=counts,
            config_a=config_a,
            config_b=config_b,
        )

    @app.route("/runs/<int:id>/delete", methods=["POST"])
    def run_delete(id):
        run = db.get_or_404(WorkflowRun, id)
        run.trashed = True
        db.session.commit()
        db_log("TRASH", "WorkflowRun", run.id, run.workflow_name)
        flash(f'Workflow run "{run.workflow_name}" moved to trash.', "success")
        return redirect(url_for("project_detail", id=run.project_id))

    @app.route("/runs/<int:id>/clear-runtime", methods=["POST"])
    def run_clear_runtime(id):
        run = db.get_or_404(WorkflowRun, id)
        run.runtime_seconds = None
        db.session.commit()
        flash("Runtime cleared.", "success")
        return redirect(url_for("run_detail", id=id))

    @app.route("/runs/batch", methods=["POST"])
    def runs_batch():
        run_ids = request.form.getlist("run_ids", type=int)
        action = request.form.get("action", "")
        value = request.form.get("value", "").strip()
        redir_kw = {
            k: request.form.get(k, "")
            for k in (
                "sort",
                "dir",
                "tag",
                "status",
                "workflow",
                "low_mapping",
                "duplicates",
                "tag_mode",
                "date_from",
                "date_to",
            )
        }
        redir_kw = {k: v for k, v in redir_kw.items() if v}

        if not run_ids:
            flash("No runs selected.", "warning")
            return redirect(url_for("runs_list", **redir_kw))

        batch_runs = WorkflowRun.query.filter(
            WorkflowRun.id.in_(run_ids),
            WorkflowRun.trashed == False,
        ).all()

        if action == "set_status":
            if value not in RUN_STATUSES:
                flash("Invalid status.", "danger")
            else:
                for run in batch_runs:
                    run.status = value
                    db_log("UPDATE", "WorkflowRun", run.id, f"batch: status → {value}")
                db.session.commit()
                label = RUN_STATUSES[value][0]
                flash(
                    f'Status set to "{label}" for {len(batch_runs)} run(s).', "success"
                )

        elif action == "add_tag":
            if not value:
                flash("No tag specified.", "warning")
            else:
                add_default_tag(value)
                updated = 0
                for run in batch_runs:
                    tags = run.tag_list
                    if value not in tags:
                        tags.append(value)
                        run.tags = ",".join(tags)
                        updated += 1
                        db_log(
                            "UPDATE",
                            "WorkflowRun",
                            run.id,
                            f"batch: tag added '{value}'",
                        )
                db.session.commit()
                flash(f'Tag "{value}" added to {updated} run(s).', "success")

        elif action == "mark_backup":
            if not value:
                flash("No backup location specified.", "warning")
            else:
                updated = 0
                for run in batch_runs:
                    blist = run.backups_list
                    if value not in [b["location"] for b in blist]:
                        blist.append({"location": value, "path": ""})
                        run.backups = json.dumps(blist)
                        updated += 1
                        db_log(
                            "UPDATE",
                            "WorkflowRun",
                            run.id,
                            f"batch: backup marked '{value}'",
                        )
                db.session.commit()
                flash(
                    f'Marked backed up at "{value}" for {updated} run(s).',
                    "success",
                )

        elif action == "remove_tags":
            updated = 0
            for run in batch_runs:
                if run.tags:
                    run.tags = ""
                    updated += 1
                    db_log("UPDATE", "WorkflowRun", run.id, "batch: all tags removed")
            db.session.commit()
            flash(f"All tags removed from {updated} run(s).", "success")

        else:
            flash("Unknown batch action.", "danger")

        return redirect(url_for("runs_list", **redir_kw))

    @app.route("/runs/<int:id>/slack", methods=["GET", "POST"])
    def run_slack(id):
        from notifier import (
            build_run_message,
            channel_from_group_name,
            send_manual_run_message,
        )

        run = db.get_or_404(WorkflowRun, id)

        # Derive channel from research group name
        channel = channel_from_group_name(run.project.researcher.group.name)

        # Workflow URL from registry
        wf_list = load_workflows()
        wf_url = next(
            (w["url"] for w in wf_list if w["name"] == run.workflow_name), None
        )

        # Extract samples for the default message
        samples = None
        for f in run.attached_files:
            if f.file_type == "config" and f.config_dict:
                extracted = extract_samples_from_config(
                    run.workflow_name, f.config_dict
                )
                if extracted:
                    samples = extracted
                    break

        if request.method == "POST":
            message = request.form.get("message", "").strip()
            override_channel = (
                request.form.get("channel", "").strip().lstrip("#") or channel
            )
            if not message:
                flash("Message cannot be empty.", "danger")
                return render_template(
                    "runs/slack_compose.html",
                    run=run,
                    channel=override_channel,
                    message=message,
                )
            ok, err = send_manual_run_message(override_channel, message)
            if ok:
                db_log(
                    "CREATE",
                    "SlackMessage",
                    run.id,
                    f"Slack message sent to #{override_channel} for run {run.workflow_name} #{run.id}",
                )
                flash(f"Message sent to #{override_channel}.", "success")
                return redirect(url_for("run_detail", id=id))
            else:
                flash(f"Failed to send Slack message: {err}", "danger")
                return render_template(
                    "runs/slack_compose.html",
                    run=run,
                    channel=override_channel,
                    message=message,
                )

        default_message = build_run_message(run, samples, wf_url=wf_url)
        return render_template(
            "runs/slack_compose.html",
            run=run,
            channel=channel,
            message=default_message,
        )
