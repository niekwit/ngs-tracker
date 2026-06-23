import os
import subprocess
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

from io import BytesIO

from flask import flash, jsonify, redirect, render_template, request, send_file, url_for
from markupsafe import Markup, escape

from config import (
    LOG_FILE,
    SETTINGS_DIR,
    WORKFLOW_SYSTEMS,
    WORKFLOWS_FILE,
    add_run_template,
    delete_run_template,
    db_log,
    load_run_templates,
    load_workflows,
    save_workflows,
)
from models import (
    Project,
    ProjectScript,
    ResearchGroup,
    Researcher,
    Sample,
    WorkflowRun,
    db,
)


def _notes_snippet(notes: str, query: str, window: int = 140) -> Markup | None:
    """Return an HTML-safe excerpt from notes around the first query match, with <mark> highlight."""
    if not notes or not query:
        return None
    lower_notes = notes.lower()
    lower_query = query.lower()
    idx = lower_notes.find(lower_query)
    if idx == -1:
        return None
    half = window // 2
    start = max(0, idx - half)
    end = min(len(notes), idx + len(query) + half)
    excerpt = notes[start:end]
    prefix = Markup("&hellip;") if start > 0 else Markup("")
    suffix = Markup("&hellip;") if end < len(notes) else Markup("")
    match_start = idx - start
    match_end = match_start + len(query)
    return (
        prefix
        + escape(excerpt[:match_start])
        + Markup("<mark>")
        + escape(excerpt[match_start:match_end])
        + Markup("</mark>")
        + escape(excerpt[match_end:])
        + suffix
    )


def register(app):
    @app.route("/workflows", methods=["GET", "POST"])
    def workflows_manage():
        if request.method == "POST":
            action = request.form.get("action")
            workflows = load_workflows()

            if action == "add":
                name = request.form.get("name", "").strip()
                url = request.form.get("url", "").strip()
                local_path = request.form.get("local_path", "").strip()
                system = request.form.get("system", "snakemake")
                if system not in WORKFLOW_SYSTEMS:
                    system = "other"
                try:
                    cutoff = float(request.form.get("mapping_rate_cutoff", "60"))
                    cutoff = max(0.0, min(100.0, cutoff))
                except ValueError:
                    cutoff = 60.0
                if not name:
                    flash("Workflow name is required.", "danger")
                elif not url and not local_path:
                    flash("Provide a GitHub URL, a local repo path, or both.", "danger")
                elif local_path and not Path(local_path).is_dir():
                    flash(f"Local path not found: {local_path}", "danger")
                elif any(w["name"] == name for w in workflows):
                    flash(f'Workflow "{name}" already exists.', "warning")
                else:
                    workflows.append(
                        {
                            "name": name,
                            "url": url,
                            "local_path": local_path,
                            "system": system,
                            "mapping_rate_cutoff": cutoff,
                        }
                    )
                    workflows.sort(key=lambda w: w["name"].lower())
                    save_workflows(workflows)
                    flash(f'Workflow "{name}" added.', "success")

            elif action == "update_cutoff":
                name = request.form.get("name", "")
                try:
                    cutoff = float(request.form.get("mapping_rate_cutoff", "60"))
                    cutoff = max(0.0, min(100.0, cutoff))
                except ValueError:
                    cutoff = 60.0
                for wf in workflows:
                    if wf["name"] == name:
                        wf["mapping_rate_cutoff"] = cutoff
                        break
                save_workflows(workflows)
                flash(f'Cutoff for "{name}" updated to {cutoff}%.', "success")

            elif action == "delete":
                name = request.form.get("name", "")
                workflows = [w for w in workflows if w["name"] != name]
                save_workflows(workflows)
                flash(f'Workflow "{name}" removed.', "success")

        return render_template(
            "workflows/manage.html",
            workflows=load_workflows(),
            templates=load_run_templates(),
            workflow_systems=WORKFLOW_SYSTEMS,
            workflows_file=str(WORKFLOWS_FILE),
        )

    @app.route("/templates/save", methods=["POST"])
    def template_save():
        run_id = request.form.get("run_id", type=int)
        name = request.form.get("name", "").strip()
        if not name:
            flash("Template name is required.", "danger")
            return redirect(url_for("run_detail", id=run_id))
        run = db.get_or_404(WorkflowRun, run_id)
        add_run_template(
            {
                "id": uuid.uuid4().hex[:8],
                "name": name,
                "workflow_name": run.workflow_name,
                "workflow_tag": run.workflow_tag or "",
                "description": run.description or "",
                "created_at": datetime.utcnow().strftime("%Y-%m-%d"),
            }
        )
        flash(f'Template "{name}" saved.', "success")
        return redirect(url_for("run_detail", id=run_id))

    @app.route("/templates/<tid>/delete", methods=["POST"])
    def template_delete(tid):
        delete_run_template(tid)
        flash("Template deleted.", "success")
        return redirect(url_for("workflows_manage"))

    @app.route("/search")
    def search():
        q = request.args.get("q", "").strip()
        if not q:
            return render_template("search.html", q="", results=None)
        like = f"%{q}%"
        results = {
            "groups": ResearchGroup.query.filter(
                ResearchGroup.trashed == False,
                ResearchGroup.name.ilike(like) | ResearchGroup.description.ilike(like),
            ).all(),
            "researchers": Researcher.query.filter(
                Researcher.trashed == False,
                Researcher.name.ilike(like) | Researcher.email.ilike(like),
            ).all(),
            "projects": Project.query.filter(
                Project.trashed == False,
                Project.name.ilike(like) | Project.description.ilike(like),
            ).all(),
            "runs": WorkflowRun.query.filter(
                WorkflowRun.trashed == False,
                WorkflowRun.workflow_name.ilike(like)
                | WorkflowRun.description.ilike(like)
                | WorkflowRun.tags.ilike(like)
                | WorkflowRun.notes.ilike(like),
            )
            .order_by(WorkflowRun.run_date.desc())
            .all(),
            "scripts": ProjectScript.query.filter(
                ProjectScript.trashed == False,
                ProjectScript.original_filename.ilike(like)
                | ProjectScript.description.ilike(like),
            ).all(),
            "samples": Sample.query.filter(
                Sample.name.ilike(like) | Sample.description.ilike(like),
            ).all(),
        }
        notes_snippets = {
            run.id: _notes_snippet(run.notes, q)
            for run in results["runs"]
            if run.notes and q.lower() in run.notes.lower()
        }
        total = sum(len(v) for v in results.values())
        return render_template(
            "search.html",
            q=q,
            results=results,
            total=total,
            notes_snippets=notes_snippets,
        )

    @app.route("/log")
    def log_viewer():
        f_user = request.args.get("user", "").strip()
        f_action = request.args.get("action", "").strip()
        f_search = request.args.get("search", "").strip()
        page = request.args.get("page", 1, type=int)
        per_page = 100

        entries = []
        if LOG_FILE.exists():
            with open(LOG_FILE) as fh:
                for line in fh:
                    line = line.rstrip("\n")
                    parts = line.split(" | ", 5)
                    if len(parts) < 5:
                        continue
                    entry = {
                        "ts": parts[0].strip(),
                        "action": parts[1].strip(),
                        "model": parts[2].strip(),
                        "record_id": parts[3].replace("id=", "").strip(),
                        "user": parts[4].replace("user=", "").strip(),
                        "detail": parts[5].strip() if len(parts) > 5 else "",
                    }
                    entries.append(entry)

        entries.reverse()

        all_users = sorted(
            {e["user"] for e in entries if e["user"] and e["user"] != "—"}
        )
        all_actions = sorted({e["action"] for e in entries})

        if f_user:
            entries = [e for e in entries if e["user"] == f_user]
        if f_action:
            entries = [e for e in entries if e["action"] == f_action]
        if f_search:
            s = f_search.lower()
            entries = [
                e
                for e in entries
                if s in e["detail"].lower()
                or s in e["model"].lower()
                or s in e["record_id"].lower()
            ]

        total = len(entries)
        pages = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, pages))
        entries = entries[(page - 1) * per_page : page * per_page]

        return render_template(
            "log.html",
            entries=entries,
            total=total,
            page=page,
            pages=pages,
            per_page=per_page,
            f_user=f_user,
            f_action=f_action,
            f_search=f_search,
            all_users=all_users,
            all_actions=all_actions,
        )

    @app.route("/workflows/xlsx-template")
    def workflow_xlsx_template():
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Protection
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            flash("openpyxl is not installed. Run: pip install openpyxl", "danger")
            return redirect(url_for("workflows_manage"))

        from config import get_default_tags

        workflow_name = request.args.get("workflow_name", "").strip()
        workflow_tag = request.args.get("workflow_tag", "").strip()

        if not workflow_name:
            flash("Select a workflow before downloading the template.", "warning")
            return redirect(url_for("workflows_manage"))

        tags = [t["name"] for t in get_default_tags()]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Run Template"

        # ── Styles ───────────────────────────────────────────────────────────
        fill_header = PatternFill("solid", fgColor="1F4E79")
        fill_locked = PatternFill("solid", fgColor="CFE2FF")
        fill_edit   = PatternFill("solid", fgColor="FFFACD")
        fill_section = PatternFill("solid", fgColor="DEE2E6")

        font_header  = Font(bold=True, color="FFFFFF", size=13)
        font_bold    = Font(bold=True)
        font_locked  = Font(bold=True, color="1F4E79")
        font_hint    = Font(italic=True, color="6C757D", size=10)

        align_center = Alignment(horizontal="center", vertical="center")
        align_wrap   = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 58

        # ── Row 1: Title ─────────────────────────────────────────────────────
        ws.merge_cells("A1:B1")
        ws["A1"] = "NGS Tracker — Workflow Run Template"
        ws["A1"].font = font_header
        ws["A1"].fill = fill_header
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 28

        # ── Row 2: Instruction ───────────────────────────────────────────────
        ws.merge_cells("A2:B2")
        ws["A2"] = (
            "Fill in the yellow cells and return this file to the bioinformatician. "
            "The Workflow and Version rows are pre-set and locked."
        )
        ws["A2"].font = font_hint
        ws["A2"].fill = PatternFill("solid", fgColor="F8F9FA")
        ws["A2"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 34

        # ── Helper: write a label cell (always locked) ───────────────────────
        def label(row, text, fill=None):
            c = ws.cell(row=row, column=1, value=text)
            c.font = font_bold
            c.fill = fill or fill_section
            c.alignment = Alignment(vertical="center")

        # ── Helper: write an editable value cell ─────────────────────────────
        def editable(row, value="", height=None, locked=False):
            c = ws.cell(row=row, column=2, value=value)
            c.fill = fill_locked if locked else fill_edit
            c.alignment = align_wrap
            c.protection = Protection(locked=locked)
            if height:
                ws.row_dimensions[row].height = height
            return c

        # ── Rows 3–4: Pre-filled, locked ─────────────────────────────────────
        label(3, "Workflow", fill=fill_locked)
        c = editable(3, workflow_name, locked=True)
        c.font = font_locked

        label(4, "Version / Release tag", fill=fill_locked)
        c = editable(4, workflow_tag or "—", locked=True)
        c.font = font_locked

        # ── Row 5: Section divider ────────────────────────────────────────────
        ws.merge_cells("A5:B5")
        ws["A5"] = "Fill in the fields below"
        ws["A5"].font = Font(bold=True, color="1F4E79", size=10)
        ws["A5"].fill = PatternFill("solid", fgColor="D0E4FF")
        ws["A5"].alignment = align_center
        ws.row_dimensions[5].height = 18

        # ── Row 6: Run Date ───────────────────────────────────────────────────
        label(6, "Run Date  (YYYY-MM-DD)")
        c = editable(6, height=22)
        c.number_format = "YYYY-MM-DD"

        # ── Row 7: Status ─────────────────────────────────────────────────────
        label(7, "Status")
        c = editable(7, "completed", height=22)
        dv_status = DataValidation(
            type="list",
            formula1='"completed,running,failed,pending"',
            showDropDown=False,
            showErrorMessage=True,
            error="Choose: completed, running, failed, or pending",
            errorTitle="Invalid status",
        )
        ws.add_data_validation(dv_status)
        dv_status.add(c)

        # ── Row 8: Short Description ──────────────────────────────────────────
        label(8, "Short Description")
        editable(8, height=26)

        # ── Row 9: Notes ──────────────────────────────────────────────────────
        label(9, "Notes")
        editable(9, height=80)

        # ── Row 10: Tags section divider ──────────────────────────────────────
        ws.merge_cells("A10:B10")
        ws["A10"] = "Tags — select Yes for each tag that applies"
        ws["A10"].font = Font(bold=True, color="1F4E79", size=10)
        ws["A10"].fill = PatternFill("solid", fgColor="D0E4FF")
        ws["A10"].alignment = align_center
        ws.row_dimensions[10].height = 18

        # ── Tag rows ──────────────────────────────────────────────────────────
        dv_yn = DataValidation(
            type="list",
            formula1='"Yes,No"',
            showDropDown=False,
            showErrorMessage=True,
            error="Choose Yes or No",
            errorTitle="Invalid value",
        )
        ws.add_data_validation(dv_yn)

        for i, tag in enumerate(tags):
            row = 11 + i
            label(row, tag)
            c = editable(row, "No", height=20)
            dv_yn.add(c)

        # ── Sheet protection (locks pre-filled cells, leaves yellow cells free) ──
        ws.protection.sheet = True
        ws.protection.password = ""

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_wf  = "".join(c if c.isalnum() or c in "-_" else "_" for c in workflow_name)
        safe_tag = "".join(c if c.isalnum() or c in "-_." else "_" for c in (workflow_tag or "no-tag"))
        filename = f"run_template_{safe_wf}_{safe_tag}.xlsx"

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/workflows/<name>/tags")
    def workflow_tags(name):
        """Return git tags (or recent commits) for a workflow with a local_path set."""
        wf = next((w for w in load_workflows() if w["name"] == name), None)
        if not wf:
            return jsonify({"error": "Workflow not found"}), 404
        local_path = wf.get("local_path", "").strip()
        if not local_path:
            return jsonify({"error": "No local path configured for this workflow"}), 400
        if not Path(local_path).is_dir():
            return jsonify({"error": f"Local path not found: {local_path}"}), 400
        try:
            tag_result = subprocess.run(
                ["git", "-C", local_path, "tag", "--sort=-version:refname"],
                capture_output=True,
                text=True,
                timeout=10,
            )
            tags = [t.strip() for t in tag_result.stdout.splitlines() if t.strip()]
            if tags:
                return jsonify({"type": "tags", "items": tags})

            # No tags — fall back to recent commits
            log_result = subprocess.run(
                [
                    "git",
                    "-C",
                    local_path,
                    "log",
                    "--oneline",
                    "-10",
                    "--format=%H\t%as\t%s",
                ],
                capture_output=True,
                text=True,
                timeout=10,
            )
            commits = []
            for line in log_result.stdout.splitlines():
                parts = line.split("\t", 2)
                if len(parts) == 3:
                    commits.append(
                        {
                            "sha": parts[0][:7],
                            "date": parts[1],
                            "subject": parts[2][:80],
                        }
                    )
            return jsonify({"type": "commits", "items": commits})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.route("/restart", methods=["POST"])
    def restart():
        def _exit():
            time.sleep(0.3)
            (SETTINGS_DIR / ".restart").touch()
            os._exit(0)

        threading.Thread(target=_exit, daemon=True).start()
        return "", 204

    @app.route("/stop", methods=["POST"])
    def stop():
        def _exit():
            time.sleep(0.3)
            os._exit(0)

        threading.Thread(target=_exit, daemon=True).start()
        return "", 204
